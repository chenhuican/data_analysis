#!/usr/bin/env python
#encoding=utf8
'''
 治疗费公医，自费计算

'''
import os,sys
from datetime import date, timedelta, datetime
from inspect import currentframe
import logging
import pprint
import MySQLdb
import _mysql
import json
import __main__
from decimal import *
from openpyxl import Workbook
from openpyxl import load_workbook
from stat import *
import time
import operator
import smtplib

getcontext().prec = 4

pp = pprint.PrettyPrinter(indent=4)
g_cur_date  = (date.today() + timedelta(0))
g_yesterday = g_cur_date - timedelta(days=1) #!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
g_export_time   = g_cur_date.strftime('%Y-%m-%d %H:%M:%S')
g_price_converter = Decimal("1.0")

###################################################################################
logging.basicConfig(format='[%(asctime)s-%(lineno)d] %(levelname)s: %(message)s', level=logging.DEBUG)

(db0_host,db0_port,db0_user,db0_passwd)    =("10.29.205.36", 3306, "user1", "password" )
(db1_host,db1_port,db1_user,db1_passwd)    =("10.30.2.118", 3306, "user1", "password")

def lineno():
    # Returns the current line number in our program.
    return currentframe().f_back.f_lineno

def close_connection(local_conn_list):
    for conn in local_conn_list:
        if conn:
            conn.close()
    local_conn_list = []

def init_deal_user_db():
    try:
        local_conn_list = []
        local_conn_list.append(MySQLdb.connect(
            host    = db0_host,
            user    = db0_user,
            passwd  = db0_passwd,
            port    = db0_port,
            db      = 'deal_user',
            charset = 'utf8'))
        local_conn_list.append(MySQLdb.connect(
            host    = db1_host,
            user    = db1_user,
            passwd  = db1_passwd,
            port    = db1_port,
            db      = 'deal_user',
            charset = 'utf8'))
        #cursor=local_conn_list.cursor(MySQLdb.cursors.Cursor)
        #pp.pprint(local_conn_list.character_set_name())
        return local_conn_list
    except MySQLdb.Error,e:
        raise RuntimeError("Ln:[%d] Mysql Error. [%d: %s]" % (lineno(), e.args[0], e.args[1]))
    except RuntimeError,e:
        if local_conn_list[0]:
             local_conn_list[0].close()
        if local_conn_list[1]:
            local_conn_list[1].close()
        raise RuntimeError(e.args[0])
    except:
        if local_conn_list[0]:
             local_conn_list[0].close()
        if local_conn_list[1]:
            local_conn_list[1].close()
        raise RuntimeError("Ln:[%d] Uncaught exception. line:[%s], desc:[%s]" % (lineno(), sys.exc_info()[2].tb_lineno, (",".join(str(x) for x in (sys.exc_info())))))

def init_deal_temporary_db():
    try:
        local_conn_list = []
        local_conn_list.append(MySQLdb.connect(
            host    = db0_host,
            user    = db0_user,
            passwd  = db0_passwd,
            port    = db0_port,
            db      = 'deal_temporary',
            charset = 'utf8'))
        return local_conn_list

    except MySQLdb.Error,e:
        raise RuntimeError("Ln:[%d] Mysql Error. [%d: %s]" % (lineno(), e.args[0], e.args[1]))
    except RuntimeError,e:
        if local_conn_list[0]:
             local_conn_list[0].close()
        raise RuntimeError(e.args[0])
    except:
        if local_conn_list[0]:
             local_conn_list[0].close()
        raise RuntimeError("Ln:[%d] Uncaught exception. line:[%s], desc:[%s]" % (lineno(), sys.exc_info()[2].tb_lineno, (",".join(str(x) for x in (sys.exc_info())))))

def init_pay_center_db():
    try:
        local_conn_list = []
        local_conn_list.append(MySQLdb.connect(
            host    = db0_host,
            user    = db0_user,
            passwd  = db0_passwd,
            port    = db0_port,
            db      = 'db_pay_center',
            charset = 'utf8'))
        local_conn_list.append(MySQLdb.connect(
            host    = db1_host,
            user    = db1_user,
            passwd  = db1_passwd,
            port    = db1_port,
            db      = 'db_pay_center',
            charset = 'utf8'))
        #cursor=local_conn_list.cursor(MySQLdb.cursors.Cursor)
        #pp.pprint(local_conn_list.character_set_name())
        return local_conn_list
    except MySQLdb.Error,e:
        raise RuntimeError("Ln:[%d] Mysql Error. [%d: %s]" % (lineno(), e.args[0], e.args[1]))
    except RuntimeError,e:
        if local_conn_list[0]:
             local_conn_list[0].close()
        if local_conn_list[1]:
            local_conn_list[1].close()
        raise RuntimeError(e.args[0])
    except:
        if local_conn_list[0]:
             local_conn_list[0].close()
        if local_conn_list[1]:
            local_conn_list[1].close()
        raise RuntimeError("Ln:[%d] Uncaught exception. line:[%s], desc:[%s]" % (lineno(), sys.exc_info()[2].tb_lineno, (",".join(str(x) for x in (sys.exc_info())))))

def load_deal_info(deal_temporary_conn, treatment_list):
    try:
        deal_info_dict = {}
        for treatment_item in treatment_list:
            deal_info_dict[treatment_item["deal_id"]] = {}

        for deal_id in deal_info_dict.keys():
            cursor = deal_temporary_conn.cursor()

            deal_sql = ((
                'SELECT '
                'Fdeal_id,'
                'Fuser_id,'
                'Fclinic_id,'
                'Freceivables_price,'
                'Fadjust_price,'
                'Fextention '
                'FROM '
                't_deal_temporary '
                'WHERE Fdeal_id = \'' + str(deal_id) + '\''))
            logging.debug(deal_sql)
            cursor.execute(deal_sql)
            deal_rows = cursor.fetchall()
            if not len(deal_rows):
                continue
            for deal_row in deal_rows:
                if not deal_row:
                    continue
                deal_info_dict[deal_id] = {
                    'Fdeal_id'           : deal_row[0],
                    'Fuser_id'           : deal_row[1],
                    'Fclinic_id'         : deal_row[2],
                    'Freceivables_price' : Decimal(deal_row[3])/g_price_converter,
                    'Fadjust_price'      : Decimal(deal_row[4])/g_price_converter,
                    'Fextention'         : deal_row[5],

                    "total_fee"          : (Decimal(deal_row[3])+Decimal(deal_row[4]))/g_price_converter,
                    "fee_discount_type"  : int(deal_row[5].split("|")[9]),
                    "fee_discount_money" : -Decimal(deal_row[5].split("|")[10])/g_price_converter,
                    "fee_discount_name"  : deal_row[5].split("|")[22],
                }

        return deal_info_dict
    except MySQLdb.Error,e:
        raise RuntimeError("Ln:[%d] Mysql Error. [%d: %s]" % (lineno(), e.args[0], e.args[1]))
    except RuntimeError,e:
        deal_temporary_conn.close()
        raise RuntimeError(e.args[0])
    except:
        deal_temporary_conn.close()
        raise RuntimeError("Ln:[%d] Uncaught exception. line:[%s], desc:[%s]" % (lineno(), sys.exc_info()[2].tb_lineno, (",".join(str(x) for x in (sys.exc_info())))))

def load_pay_info(pay_center_conn, deal_dict):
    try:
        for deal_id,deal_item in deal_dict.iteritems():
            clinic_id = deal_item["Fclinic_id"]
            cursor = pay_center_conn[clinic_id%2].cursor()
            suffix = '%02d' % (clinic_id % 100)

            pay_center_sql = (('SELECT '
                'Fpay_type,Fdeal_id,Fpay_name,SUM(Fpay_fee)'
                'FROM '
                't_pay_item_' + suffix + ' '
                'WHERE Fdeal_id = \'' + str(deal_id) + '\' '
                'AND Fpay_scene <= 4 '
                'GROUP BY Fpay_type;'))
            logging.debug(pay_center_sql)
            cursor.execute(pay_center_sql)
            pay_center_rows = cursor.fetchall()
            if not len(pay_center_rows):
                continue
            for pay_center_row in pay_center_rows:
                if not pay_center_row:
                    continue
                pay_type = int(pay_center_row[0])
                deal_id  = int(pay_center_row[1])

                if deal_id not in deal_dict.keys():
                    raise RumtimeError("Ln:[%d] DealId:[%s] not exist." % (lineno(),deal_id))
                if "pay_info" not in deal_dict[deal_id].keys():
                    deal_dict[deal_id]["pay_info"] = {}

                deal_dict[deal_id]["pay_info"][pay_type] = {
                        'Fpay_type' : pay_center_row[0],
                        'Fdeal_id'  : pay_center_row[1],
                        'Fpay_name' : pay_center_row[2],
                        'Fpay_fee'  : Decimal(pay_center_row[3])/g_price_converter,
                }

        return deal_dict
    except MySQLdb.Error,e:
        raise RuntimeError("Ln:[%d] Mysql Error. [%d: %s]" % (lineno(), e.args[0], e.args[1]))
    except RuntimeError,e:
        close_connection(local_conn_list)
        raise RuntimeError(e.args[0])
    except:
        close_connection(local_conn_list)
        raise RuntimeError("Ln:[%d] Uncaught exception. line:[%s], desc:[%s]" % (lineno(), sys.exc_info()[2].tb_lineno, (",".join(str(x) for x in (sys.exc_info())))))

def load_treatment_list_from_excel(filename):
    try:
        treatment_list = []
        wb = load_workbook(filename=filename, read_only=True)
        sheet = wb.active
        #print sheet.title
        first_row = True
        index = 1
        for row in sheet.rows:
            if first_row:
                first_row = False
                continue

            index+=1
            print index, row[10].value, row[12].value
            treatment_list.append({
                "date"             : row[0].value.strip(),
                "city_id"          : int(row[1].value),
                "cloud_recipe_id"  : int(row[2].value),
                "deal_id"          : int(row[3].value),
                "clinic_id"        : int(row[4].value),
                "clinic_name"      : row[5].value.strip(),
                "department_code"  : int(row[6].value),
                "department_name"  : row[7].value.strip(),
                "doctor_id"        : int(row[8].value),
                "doctor_name"      : row[9].value.strip(),
                "item_code"        : row[10].value.strip(),
                "item_name"        : row[11].value.strip(),
                "item_unit"        : row[12].value,
                "item_spec"        : row[13].value,
                "class1"           : row[14].value.strip(),
                "class2"           : row[15].value.strip(),
                "class3"           : row[16].value.strip(),
                "basic_class"      : int(row[17].value),
                "basic_class_name" : row[18].value.strip(),
                "amount"           : int(row[19].value),
                "price"            : Decimal(str(row[20].value).strip())*Decimal(10000),
                })
            #break

        return treatment_list
    except RuntimeError,e:
        raise RuntimeError(e.args[0])
    except:
        raise RuntimeError("Ln:[%d] Uncaught exception. line:[%s], desc:[%s]" % (lineno(), sys.exc_info()[2].tb_lineno, (",".join(str(x) for x in (sys.exc_info())))))

def calculate_treament_price(treatment_list, deal_dict):
    try:
        for deal_id, deal_item in deal_dict.iteritems():
            deal_item["pgf"] = Decimal('0')     #省公医
            deal_item["cgf"] = Decimal('0')     #市公医

            if deal_item["fee_discount_type"] >= 28 and deal_item["fee_discount_type"] <= 55:
                deal_item["pgf"] += deal_item['fee_discount_money']

            for pay_type,pay_info in deal_item["pay_info"].iteritems():
                if pay_type == 6:
                    deal_item["pgf"] += pay_info["Fpay_fee"]
                elif pay_type == 57:
                    deal_item["cgf"] += pay_info["Fpay_fee"]

            deal_item["self_fee"] = deal_item["total_fee"] - deal_item["pgf"] - deal_item["cgf"]

        for treatment in treatment_list:
            deal_id = int(treatment["deal_id"])
            treatment["total_fee"] = deal_dict[deal_id]["total_fee"]
            treatment["pgf"]       = Decimal('0')
            treatment["cgf"]       = Decimal('0')

            if deal_dict[deal_id]["total_fee"] == 0:
                continue

            treatment["pgf"]       = deal_dict[deal_id]["pgf"] * (treatment["price"]/deal_dict[deal_id]["total_fee"])
            treatment["cgf"]       = deal_dict[deal_id]["cgf"] * (treatment["price"]/deal_dict[deal_id]["total_fee"])

    except RuntimeError,e:
        raise RuntimeError(e.args[0])
    except:
        raise RuntimeError("Ln:[%d] Uncaught exception. line:[%s], desc:[%s]" % (lineno(), sys.exc_info()[2].tb_lineno, (",".join(str(x) for x in (sys.exc_info())))))

if __name__ == "__main__":
    try:
        os.chdir(os.path.dirname(__file__))
        if(len(sys.argv) < 2):
            pp.pprint(sys.argv)
            pp.pprint("Usage:%s treatment_price_file.xlsx " % (sys.argv[0]))
            sys.exit(1)

        filename = sys.argv[1]
        treatment_list = load_treatment_list_from_excel(filename)
        #pp.pprint(treatment_list)

        deal_temporary_conn = init_deal_temporary_db()[0]
        deal_dict = load_deal_info(deal_temporary_conn, treatment_list)
        #pp.pprint(deal_dict)

        pay_center_conn = init_pay_center_db()
        pay_info_dict = load_pay_info(pay_center_conn, deal_dict)
        #pp.pprint(pay_info_dict)

        calculate_treament_price(treatment_list, deal_dict);
        #pp.pprint(pay_info_dict)
        #pp.pprint(treatment_list)

        result_book = Workbook()
        result_book.remove_sheet(result_book.active)
        deal_sheet      = result_book.create_sheet(title=u"订单数据明细", index=2)
        deal_sheet['A1'] = '订单Id'
        deal_sheet['B1'] = '用户Id'
        deal_sheet['C1'] = '门店Id'
        deal_sheet['D1'] = '总金额'
        deal_sheet['E1'] = '省公医'
        deal_sheet['F1'] = '市公医'
        deal_sheet['G1'] = '自付金额'   #self_fee

        i = 2
        for deal_id,deal_item in deal_dict.iteritems():
            deal_sheet['A%s'%i] = deal_item['Fdeal_id']
            deal_sheet['B%s'%i] = deal_item['Fuser_id']
            deal_sheet['C%s'%i] = deal_item['Fclinic_id']
            deal_sheet['D%s'%i] = deal_item['total_fee']/Decimal(10000)
            deal_sheet['E%s'%i] = deal_item['pgf']/Decimal(10000)
            deal_sheet['F%s'%i] = deal_item['cgf']/Decimal(10000)
            deal_sheet['G%s'%i] = deal_item['self_fee']/Decimal(10000)
            i += 1

        treatment_sheet = result_book.create_sheet(title=u"治疗数据明细", index=3)
        treatment_sheet['A1'] = '日期'
        treatment_sheet['B1'] = 'City_Id'
        treatment_sheet['C1'] = 'Cloud_recipe_Id'
        treatment_sheet['D1'] = 'Deal_Id'
        treatment_sheet['E1'] = 'Shop_Id'
        treatment_sheet['F1'] = 'Shop_Name'
        treatment_sheet['G1'] = 'Dept_Code'
        treatment_sheet['H1'] = 'Dept_Name'
        treatment_sheet['I1'] = '医生工号'
        treatment_sheet['J1'] = '医生项姓名'
        treatment_sheet['K1'] = '项目编码'
        treatment_sheet['L1'] = '项目名称'
        treatment_sheet['M1'] = '单位'
        treatment_sheet['N1'] = '规格'
        treatment_sheet['O1'] = '一级分类'
        treatment_sheet['P1'] = '二级分类'
        treatment_sheet['Q1'] = '三级分类'
        treatment_sheet['R1'] = '基础分类代码'
        treatment_sheet['S1'] = '基础分类'
        treatment_sheet['T1'] = '数量'
        treatment_sheet['U1'] = '零售总金额'
        treatment_sheet['V1'] = '省公费'
        treatment_sheet['W1'] = '市公费'

        i = 2
        for treatment in treatment_list:
            treatment_sheet['A%s'%i] = treatment["date"]
            treatment_sheet['B%s'%i] = treatment["city_id"]
            treatment_sheet['C%s'%i] = treatment["cloud_recipe_id"]
            treatment_sheet['D%s'%i] = treatment["deal_id"]
            treatment_sheet['E%s'%i] = treatment["clinic_id"]
            treatment_sheet['F%s'%i] = treatment["clinic_name"]
            treatment_sheet['G%s'%i] = treatment["department_code"]
            treatment_sheet['H%s'%i] = treatment["department_name"]
            treatment_sheet['I%s'%i] = treatment["doctor_id"]
            treatment_sheet['J%s'%i] = treatment["doctor_name"]
            treatment_sheet['K%s'%i] = treatment["item_code"]
            treatment_sheet['L%s'%i] = treatment["item_name"]
            treatment_sheet['M%s'%i] = treatment["item_unit"]
            treatment_sheet['N%s'%i] = treatment["item_spec"]
            treatment_sheet['O%s'%i] = treatment["class1"]
            treatment_sheet['P%s'%i] = treatment["class2"]
            treatment_sheet['Q%s'%i] = treatment["class3"]
            treatment_sheet['R%s'%i] = treatment["basic_class"]
            treatment_sheet['S%s'%i] = treatment["basic_class_name"]
            treatment_sheet['T%s'%i] = treatment["amount"]
            treatment_sheet['U%s'%i] = treatment["price"]/Decimal(10000)
            treatment_sheet['V%s'%i] = treatment["pgf"]/Decimal(10000)
            treatment_sheet['W%s'%i] = treatment["cgf"]/Decimal(10000)

            i += 1


        fee_dict = {}
        for treatment in treatment_list:
            date        = treatment['date']
            doctor_id   = treatment['doctor_id']
            doctor_name = treatment['doctor_name']
            clinic_name = treatment['clinic_name']

            if date not in fee_dict.keys():
                fee_dict[date] = {}
            if clinic_name not in fee_dict[date]:
                fee_dict[date][clinic_name] = {}
            if doctor_id not in fee_dict[date][clinic_name]:
                fee_dict[date][clinic_name][doctor_id] = {
                        "doctor_id"             : doctor_id,
                        "doctor_name"           : doctor_name,
                        "clinic_name"           : clinic_name,
                        "total_fee"             : Decimal(0),
                        "pgf"                   : Decimal(0),
                        "cgf"                   : Decimal(0),
                        "self_fee"              : Decimal(0),
                        "deal_list"             : [],
                        }

            if treatment["deal_id"] not in fee_dict[date][clinic_name][doctor_id]["deal_list"]:
                fee_dict[date][clinic_name][doctor_id]["deal_list"].append(treatment["deal_id"])
                fee_dict[date][clinic_name][doctor_id]["self_fee"] += deal_dict[treatment["deal_id"]]["self_fee"]

            if treatment["class2"] != u"治疗费":
                continue

            #print date,treatment["class2"]
            fee_dict[date][clinic_name][doctor_id]["total_fee"] += treatment["price"]
            fee_dict[date][clinic_name][doctor_id]["pgf"]       += treatment["pgf"]
            fee_dict[date][clinic_name][doctor_id]["cgf"]       += treatment["cgf"]


        pp.pprint(fee_dict)
        fee_sheet = result_book.create_sheet(title=u"医生治疗费业绩统计", index=0)
        fee_sheet['A1'] = '门店名称'
        fee_sheet['B1'] = '医生工号'
        fee_sheet['C1'] = '医生姓名'
        fee_sheet['D1'] = '公医治疗费(省市)'
        fee_sheet['E1'] = '自费治疗费'
        fee_sheet['F1'] = '自付金额'
        fee_sheet['G1'] = '日期'
        fee_sheet['H1'] = '治疗费总金额'
        fee_sheet['I1'] = '省公医治疗费'
        fee_sheet['J1'] = '市公医治疗费'

        i = 2
        for date, date_item in fee_dict.iteritems():
            for clinic_name, clinic_item in date_item.iteritems():
                for doctor_id, doctor_item in clinic_item.iteritems():
                    fee_sheet['A%s'%i] = clinic_name
                    fee_sheet['B%s'%i] = doctor_item["doctor_id"]
                    fee_sheet['C%s'%i] = doctor_item["doctor_name"]
                    fee_sheet['D%s'%i] = (doctor_item["pgf"] + doctor_item["cgf"])/Decimal(10000)
                    fee_sheet['E%s'%i] = (doctor_item["total_fee"] - (doctor_item["pgf"] + doctor_item["cgf"]))/Decimal(10000)
                    fee_sheet['F%s'%i] = (doctor_item["self_fee"])/Decimal(10000)
                    fee_sheet['G%s'%i] = date
                    fee_sheet['H%s'%i] = doctor_item["total_fee"]/Decimal(10000)
                    fee_sheet['I%s'%i] = doctor_item["pgf"]/Decimal(10000)
                    fee_sheet['J%s'%i] = doctor_item["cgf"]/Decimal(10000)
                    i += 1

        result_filename = '%s.result.xlsx' % (filename.replace('.xlsx',''))
        result_book.save(result_filename)

    except RuntimeError,e:
        error_message = str("%s" % e.args[0])
        logging.error(error_message)
        sys.exit()
