#!/usr/bin/env python
#coding:utf-8
'''
 考勤数据处理
 author:huican.chen
'''
import os
import sys
import time
import copy
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, colors, Alignment
import xlrd
import threading
import pandas as pd


class Convert_Excel():

	def __init__(self, filename):
		self.path = os.path.abspath('.')
		self.days = time.strftime("%Y%m%d%H%M%S", time.localtime())
		self.listday = []
		self.dicday = {}
		self.workname = set()
		self.filename = self.xls_to_xlsx(filename)

	def read_excel(self):
		if self.filename is None:
			print('Error:文件为空，退出！')
			sys.exit()
			
		try:
			wb = load_workbook(os.path.join(self.path, self.filename))
			sheet = wb.get_active_sheet()
			colnums = sheet.max_column
			rownums = sheet.max_row
			parse_cols = [3,4,7,8,9]
			for k in range(2, rownums+1):
				ws_day = sheet.cell(row=k, column=4).value
				self.dicday[ws_day] = []
			
			name_list = []
			for rows in range(2, rownums+1):
				#for col in parse_cols:
				w_name = sheet.cell(row=rows, column=3).value
				name_list.append(w_name)
				w_day = sheet.cell(row=rows, column=4).value
				w_uptime   = sheet.cell(row=rows, column=7).value
				w_downtime = sheet.cell(row=rows, column=8).value 
				w_worktime = sheet.cell(row=rows, column=9).value
				dict_day = {
					u"姓名":w_name,
					u"上班打卡时间":w_uptime,
					u"下班打卡时间":w_downtime,
					u"工作时间":w_worktime}

				self.dicday[w_day].append(dict_day)

				# print(w_name,w_day,w_uptime,w_downtime,w_worktime)
			self.listday.append(self.dicday)
			self.listday.sort()
			# 去重获取所有员工姓名
			self.workname = set(name_list)
			print(self.workname)

		except Exception, e:
			print(str(e))
		
		return True
	
	def write_excel(self):
		wb = Workbook()
		ws = wb.create_sheet(title='result', index=0)
		bold_itatic_14_font = Font(name='Arial', size=14, italic=True, color=colors.RED, bold=True)
		bold_12_font = Font(name='Arial', size=11, italic=False, color=colors.RED, bold=True)
		c = 0 
		letters = self.general_cell()
		list_a = [u"上班打卡时间", u"下班打卡时间", u"工作时间"]
		list_title = [k for i in range(len(self.dicday)*3) for k in list_a]
		
		# 工作日列表
		work_day = sorted(self.dicday.keys(), reverse=False)

		for col_name in work_day:
			c += 1
			if c == 1:
				st, et = 1, 3
			else:
				st, et = st+3, et+3
			#print('[合并单元格]:(st:%s, et:%s)'%(st, et))
			ws.merge_cells('{}:{}'.format(letters[st]+'1', letters[et]+'1'))
			ws.cell(row=1, column=(2 if st==0 else st+1)).font = bold_itatic_14_font
			ws.cell(row=1, column=(2 if st==0 else st+1)).alignment = Alignment(horizontal='center', vertical='center')
			ws.cell(row=1, column=(2 if st==0 else st+1)).value = str(col_name)
			# print(c, col_name)
		
		# 写副标题
		for i in range(1, len(self.dicday)*3+1):
			ws.cell(row=2, column=i+1).value = list_title[i-1]

		
		workname_list = list(self.workname)
		for r in range(len(workname_list)):
			col = 1
			for days in work_day:
				for kv in range(len(list_a)):
					col += 1
					ws.cell(row=r+3, column=1).value = workname_list[r]
					values = self.get_info(days, workname_list[r], kv, list_a) 
					ws.cell(row=r+3, column=col).value = values 
					#print([j.get(list_a[kv],"N") for j in self.dicday[days] for k,v in j.items() if k==u"姓名" and v==workname_list[r]])
				
		wb.save('gst_result_{}@360gst.xlsx'.format(str(self.days)))
	
	def convert_excel(self):
		pass

	
	def xls_to_xlsx(self, filename):
		''' old xls convert to xlsx.
		    避免报错：No CODEPAGE record, no encoding_override: will use 'ascii'
		    采用先 xlrd 读取转换
		'''
		files = os.path.join(self.path, filename)
		print 'files: {}'.format(files)
		suffix = filename.split('.')[1]
		file_name = str(filename.split('.')[0])
		if suffix=='xls':
			content = xlrd.open_workbook(filename=files, encoding_override='gbk')
			df = pd.read_excel(content, engine='xlrd')
			# df = pd.read_excel(files)
			files = file_name+'.xlsx'
			df.to_excel(files, index=False)

		return files 

	def get_info(self, days, work_name, kv, list_a):
		values = ''
		for j in self.dicday[days]:
			for k,v in j.items():
				if k==u"姓名" and v==work_name:
					values = j.get(list_a[kv],"")
		# print work_name,values
		return values

	def general_cell(self):
		'''生成合并单元格列表'''
		letter_a = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
		lists = ['A','B','C','D']
		self.letters = copy.copy(letter_a)
		for k in lists:
			for v in letter_a:
				self.letters.append(str(k)+str(v))
		return self.letters


if __name__ == "__main__":
	if len(sys.argv) >= 2:
		filename = sys.argv[1]
		if filename.split('.')[1] not in ['xls','xlsx']:
			print("{} is not a excel file, please input excel file !\n".format(filename))
			sys.exit(1)
	else:
		print("Usage: %s filename "%(sys.argv[0]))
		sys.exit(1)

	c_excels = Convert_Excel(filename)
	c_excels.read_excel()
	c_excels.write_excel()
