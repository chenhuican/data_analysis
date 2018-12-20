# coding:utf-8
'''
    财务进销存先进先出处理逻辑
    author:huican.chen
'''
import os
import sys
import time
import copy
from inspect import currentframe
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, colors, Alignment
import xlrd
import pandas as pd


class ConvertExcel(object):

    def __init__(self, filename):
        self.path = os.path.abspath('.')
        self.days = time.strftime("%Y%m%d%H%M%S", time.localtime())
        self.listday = []
        self.dicday = {}
        self.workname = set()
        self.filename = self.xls_to_xlsx(filename)

    def read_financial_excel(self):
        if self.filename is None:
            print('Error:文件为空，退出！')
            sys.exit()
        try:
            self.temp_file = os.path.join(self.path, self.filename)
            wb = load_workbook(os.path.join(self.path, self.filename))
            sheet_list = wb.sheetnames
            sheetb = wb[sheet_list[0]]
            sheeta = wb[sheet_list[1]]
            b_rownums = sheetb.max_row
            a_rownums = sheeta.max_row
            # sheet = wb.get_active_sheet()
	    print b_rownums,a_rownums
            for i in range(2, b_rownums + 1):
                money_01 = 0  # 含税成本
                money_02 = 0  # 含税收入
                money_03 = 0  # 不含税成本
                money_04 = 0  # 不含税收入
                total_amount = 0
                shopid = int(sheetb['A%s' % i].value)
                shopname = sheetb['B%s' % i].value
                itemcode = str(sheetb['C%s' % i].value)
                bmount = float(sheetb['F%s' % i].value)
		print shopid,shopname,itemcode
                for k in range(2, a_rownums + 1):
		    print sheeta['B%s' % k].value,sheeta['D%s' % k].value
                    if shopid == int(sheeta['B%s' % k].value) and itemcode == str(sheeta['D%s' % k].value):
		        print sheeta['B%s' % k].value,sheeta['D%s' % k].value
                        price01 = float(sheeta['H%s' % k].value) / float(sheeta['G%s' % k].value)
                        price02 = float(sheeta['I%s' % k].value) / float(sheeta['G%s' % k].value)
                        price03 = float(sheeta['J%s' % k].value) / float(sheeta['G%s' % k].value)
                        price04 = float(sheeta['K%s' % k].value) / float(sheeta['G%s' % k].value)
                        amount = float(sheeta['G%s' % k].value)
                        if bmount <= amount:
                            money_01 += price01 * bmount
                            money_02 += price02 * bmount
                            money_03 += price03 * bmount
                            money_04 += price04 * bmount
                            break
                        else:
                            total_amount += amount
                            if (bmount - total_amount) >= 0:
                                money_01 += price01 * amount
                                money_02 += price02 * amount
                                money_03 += price03 * amount
                                money_04 += price04 * amount
		    		print money_01,money_02,money_03,money_04
                            else:
                                money_01 += price01 * (bmount - total_amount+amount)
                                money_02 += price02 * (bmount - total_amount+amount)
                                money_03 += price03 * (bmount - total_amount+amount)
                                money_04 += price04 * (bmount - total_amount+amount)
                                break
                sheetb['G%s' % i] = money_01
                sheetb['H%s' % i] = money_02
                sheetb['I%s' % i] = money_03
                sheetb['J%s' % i] = money_04

            wb.save(filename='new'+ self.filename)

        except Exception as e:
            print('line: '+str(currentframe().f_back.f_lineno) + str(e))

    def convert_excel(self):
        self.read_financial_excel()

    def xls_to_xlsx(self, filename):
        ''' old xls convert to xlsx.
            避免报错：No CODEPAGE record, no encoding_override: will use 'ascii'
            采用先 xlrd 读取转换
        '''
        print("Convert file is: %s" % (filename))
        files = filename.split("/")[-1]
        print('files: %s' % files)
        suffix = files.split('.')[1]
        file_name = str(files.split('.')[0])
        if suffix == 'xls':
            content = xlrd.open_workbook(filename=filename, encoding_override='gbk')
            df = pd.read_excel(content, engine='xlrd')
            # df = pd.read_excel(files)
            files = file_name + '.xlsx'
            df.to_excel(files, index=False)

        return files


if __name__ == "__main__":
    if len(sys.argv) >= 2:
        filename = sys.argv[1]
        if filename.split('.')[1] not in ['xls', 'xlsx']:
            print("{} is not a excel file, please input excel file !\n".format(filename))
            sys.exit(1)
    else:
        print("Usage: %s filename " % (sys.argv[0]))
        sys.exit(1)

    c_excels = ConvertExcel(filename)
    c_excels.convert_excel()
